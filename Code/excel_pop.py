from openpyxl import load_workbook
import datetime


def find_date_cells(filename, target_date):
  """
  This function takes an xlsx filename, a target date (Python date object), 
  and returns a list of tuples containing the row/column headings for all cells 
  containing the target date.

  Args:
      filename: Path to the xlsx file.
      target_date: The date to search for (can be a Python date object).

  Returns:
      A list of tuples (row_heading, col_heading) for cells containing the date.
  """

  # Load the workbook from the specified filename
  wb = load_workbook(filename, read_only=True)
  # List to store cells containing the target date
  found_cells = []
  # Get the first sheet by default (modify if needed)
  for sheet_name in wb.sheetnames:
    sheet = wb[sheet_name]
    found_per_sheet = []

    # Iterate through all rows (skip the first row assumed to be headings)
    for row in sheet.iter_rows(min_row=1):
      # Indent the code block inside the for loop
      for cell in row:
          if cell.value is not None:  # Check if cell value is not None
              if isinstance(cell.value, str) and cell.value.startswith('='):
                  if cell.value ==  target_date:
                      row_heading = sheet.cell(row=cell.row, column=1).value  # Assuming row heading in column 1
                      col_heading = sheet.cell(row=1, column=cell.column).value  # Assuming column heading in row 1

                      # Store results for the sheet
                      found_per_sheet.append((row_heading, col_heading))  
                      # ... (existing code to handle formula - modify if needed)
                      pass  # Logic for handling formulas (placeholder)
    for row_heading, col_heading in found_per_sheet:
      # Store results for all sheets
      found_cells.append((sheet_name, row_heading, col_heading))  

  # Return the list of cells containing the target date
  return found_cells

def convert_to_date_formula(day, month, year):
  # Format day, month, and year with leading zeros (optional for aesthetics)
  day_str = f"{day:02d}"
  month_str = f"{month:02d}"
  # Remove all zeros from the beginning of the month and day
  month_str = month_str.lstrip("0")
  day_str = day_str.lstrip("0")
  
  year_str = str(year)

  # Construct the formula string
  formula_string = f"=DATE({year_str},{month_str},{day_str})"

  return formula_string

# Example usage
day = 2
month = 5
year = 2024

target_date = convert_to_date_formula(day, month, year)
print(target_date)

# Example usage
filename = "C:/Users/aayus/Desktop/Bio.xlsx"

try:
  found_cells = find_date_cells(filename, target_date)
  if found_cells:
      #convert target date to string for comparison
      formula_string = target_date

      # Remove leading characters (assuming "=DATE(" prefix)
      date_string = formula_string[6:]

      # Split the string into year, month, day
      year, month, day = date_string.strip("()").split(",")

      # Convert month and day to integers (optional, you can keep them as strings)
      month = int(month.lstrip("0"))
      day = int(day.lstrip("0"))

      # Format the date string with leading zeros removed for month and day
      target_date = f"{year}-{month:02d}-{day:02d}"
      print(f"To-Do List of {target_date}")
      x=""
      for cell in found_cells:
        sheet_name, row_heading, col_heading = cell
        if(sheet_name != x):
          print("")
        print(f"{sheet_name}\t => {col_heading} Of {row_heading}")
        
        x = sheet_name
  else:
      print(f"No cells found containing the date: {target_date}")
except ValueError as e:
  print(e)
